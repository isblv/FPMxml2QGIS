import os
from qgis.PyQt.QtWidgets import QAction, QFileDialog, QToolBar
from qgis.core import QgsVectorLayer, QgsField, QgsFeature, QgsGeometry, QgsPointXY, QgsProject
from PyQt5.QtWidgets import QAction, QCheckBox
#from PyQt5.QtWidgets import QCheckBox
from PyQt5.QtCore import QVariant, QSettings
#from PyQt5.QtGui import QIcon
import xml.etree.ElementTree as ET
import tempfile
import shutil
import subprocess
import inspect
import openpyxl
import math
from datetime import datetime
from openpyxl import load_workbook

# Получение пути папки, содержащую текущий плагин
cmd_folder = os.path.split(inspect.getfile(inspect.currentframe()))[0]
# Временное название
_name = 'XML to GeoCSV'

class XML2QGISPlugin:
    def __init__(self, iface):
        self.iface = iface
        self.layers = []

        # Создание временной директории
        self.temp_dir = tempfile.mkdtemp()

        self.checkboxUseAdditionalInfoAction = None  # Добавляем атрибут checkboxUseAdditionalInfoAction

        # Путь к файлу Excel с лесными стратами (+ районами, зонами ЛПУ и т.д.) (Приложение 1)
        self.excel_file_path = os.path.join(cmd_folder, "Приложение 1.xlsx")
        
    def initGui(self):
        # Создание панели инструментов
        self.toolbar = self.iface.addToolBar(_name+" панель инструментов")

        #Создание инструментов для открытия XML-файлов, папок и архивов
        self.actionOpenXmlFiles = QAction('Файл', self.iface.mainWindow())
        self.actionOpenFolders = QAction('Папка', self.iface.mainWindow())
        self.actionOpenArchives = QAction('Архив', self.iface.mainWindow())

        self.actionOpenXmlFiles.triggered.connect(self.open_xml_files)
        self.actionOpenFolders.triggered.connect(self.open_folders)
        self.actionOpenArchives.triggered.connect(self.open_archives)

        self.toolbar.addAction(self.actionOpenXmlFiles)
        self.toolbar.addAction(self.actionOpenFolders)
        self.toolbar.addAction(self.actionOpenArchives)

        self.actionOpenExcel = QAction('Страты', self.iface.mainWindow())
        self.actionOpenExcel.triggered.connect(self.load_excel_data_to_map)
        self.toolbar.addAction(self.actionOpenExcel)

        self.iface.addPluginToMenu("&"+_name, self.actionOpenXmlFiles)
        self.iface.addPluginToMenu("&"+_name, self.actionOpenFolders)
        self.iface.addPluginToMenu("&"+_name, self.actionOpenArchives)
        self.iface.addPluginToMenu("&"+_name, self.actionOpenExcel)


        # Создать действие для использования дополнительной информации
        self.checkboxUseAdditionalInfoAction = QAction("Подеревная", self.iface.mainWindow())
        self.checkboxUseAdditionalInfoAction.setCheckable(True)
        self.checkboxUseAdditionalInfoAction.setChecked(False)  # Начальное значение чекбокса

        # Привязать действие к слоту, который будет вызываться при изменении состояния чекбокса
        self.checkboxUseAdditionalInfoAction.triggered.connect(self.onCheckboxStateChanged)

        # Добавить действие в панель инструментов
        self.toolbar.addAction(self.checkboxUseAdditionalInfoAction)

        # Добавить действие в меню плагина
        self.iface.addPluginToMenu("&" + _name, self.checkboxUseAdditionalInfoAction)

        self.iface.messageBar().pushMessage('Инструменты плагина ' + _name + ' загружены в панель инструментов')


    def unload(self):        
        # Удаление действий из панели инструментов при выгрузке плагина
        self.iface.removePluginMenu("&"+_name, self.actionOpenXmlFiles)
        self.iface.removePluginMenu("&"+_name, self.actionOpenFolders)
        self.iface.removePluginMenu("&"+_name, self.actionOpenArchives)
        self.iface.removePluginMenu("&"+_name, self.actionOpenExcel)
        '''self.iface.removePluginMenu("&"+_name, self.checkboxUseAdditionalInfo)

        # Удалить чекбокс из панели инструментов
        self.toolbar.removeAction(self.checkboxUseAdditionalInfoAction)

        # Удаление действий из панели инструментов при выгрузке плагина
        self.toolbar.removeAction(self.actionOpenXmlFiles)
        self.toolbar.removeAction(self.actionOpenFolders)
        self.toolbar.removeAction(self.actionOpenArchives)
        self.toolbar.removeAction(self.actionOpenExcel)
        
        # Очистка панели инструментов
        self.toolbar.clear()

        # Удаление чекбокса
        self.toolbar.removeAction(self.checkboxUseAdditionalInfoAction)

        # Очистить указатели на элементы интерфейса
        self.actionOpenXmlFiles = None
        self.actionOpenFolders = None
        self.actionOpenArchives = None
        self.actionOpenExcel = None
        self.checkboxUseAdditionalInfo = None'''

        # Удалить действие из меню плагина при выгрузке плагина
        self.iface.removePluginMenu("&" + _name, self.checkboxUseAdditionalInfoAction)
        # Очистить панель инструментов
        del self.checkboxUseAdditionalInfoAction
        self.toolbar.clear()

    def onCheckboxStateChanged(self):
        # Обработка изменения состояния чекбокса
        if self.checkboxUseAdditionalInfoAction.isChecked():
            # Чекбокс отмечен
            print("Использование дополнительной информации включено")
        else:
            # Чекбокс не отмечен
            print("Использование дополнительной информации выключено")

    def open_xml_files(self):
        # Открыть диалоговое окно для выбора XML-файлов
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("XML (*.xml)")
        if file_dialog.exec_():
            # Получить список выбранных файлов
            xml_files = file_dialog.selectedFiles()
            # Обработать выбранные файлы
            self.process_files(xml_files)

    def open_folders(self):
        # Открыть диалоговое окно для выбора папки
        folder_dialog = QFileDialog()
        folder_dialog.setFileMode(QFileDialog.Directory)
        folder_dialog.setOption(QFileDialog.ShowDirsOnly)
        folder_dialog.setAcceptMode(QFileDialog.AcceptOpen)
        if folder_dialog.exec_():
            # Получить путь к выбранной папке
            folder_path = folder_dialog.selectedFiles()[0]
            # Проверить, выбран ли архив или папка
            if folder_path.endswith('.rar') or folder_path.endswith('.7z') or folder_path.endswith('.zip'):
                # Если выбран архив, создать временную папку
                self.temp_dir = tempfile.mkdtemp()
                # Обработать архив
                xml_files = self.get_xml_files_in_archive(folder_path)
                # Удалить временную папку после обработки
                shutil.rmtree(self.temp_dir)
            else:
                # Если выбрана папка или файл, обработать её содержимое без использования временной папки
                if os.path.isdir(folder_path):
                    xml_files = self.get_xml_files_in_folder(folder_path)
                else:
                    xml_files = [folder_path]
            # Обработать XML-файлы
            self.process_files(xml_files)

    def open_archives(self):
        # Открыть диалоговое окно для выбора архивов
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("Архив (*.zip *.rar *.7z)")
        if file_dialog.exec_():
            # Получить список выбранных архивов
            archive_files = file_dialog.selectedFiles()
            for archive_file in archive_files:
                # Получить список XML-файлов в архиве
                xml_files = self.get_xml_files_in_archive(archive_file)
                # Обработать XML-файлы
                self.process_files(xml_files)

    def process_files(self, xml_files):
        if not xml_files:
            return
        
        # Создать новый слой (новый метод обработки в зависимости от значения чекбокса)
        if self.checkboxUseAdditionalInfoAction.isChecked():
            layer = self.create_layer_with_additional_info()
            if layer is None:
                print("Не удалось создать слой")
                return
            # Обработать каждый XML-файл и добавить объекты на слой
            for xml_file in xml_files:
                print(f"Обработка файла: {xml_file}")
                self.process_xml_file_with_additional_info(xml_file, layer)
        else:
            layer = self.create_layer()
            if layer is None:
                print("Не удалось создать слой")
                return
            # Обработать каждый XML-файл и добавить объекты на слой
            for xml_file in xml_files:
                print(f"Обработка файла: {xml_file}")
                self.process_xml_file(xml_file, layer)

        # Создать новый слой
        '''layer = self.create_layer()
        if layer is None:
            print("Не удалось создать слой")
            return

        # Обработать каждый XML-файл и добавить объекты на слой
        for xml_file in xml_files:
            self.process_xml_file(xml_file, layer)'''

        # Добавить слой в проект
        self.layers.append(layer)
        QgsProject.instance().addMapLayer(layer)

    def create_layer(self):
        # Создать новый векторный слой
        layer = QgsVectorLayer("Point", "XML ППН", "memory")
        if not layer.isValid():
            return None
        provider = layer.dataProvider()
        provider.addAttributes([QgsField("xml_path", QVariant.String),
                                QgsField("path_area", QVariant.Double),
                                QgsField("tax_info", QVariant.String),
                                QgsField("strata", QVariant.String),
                                QgsField("tree_count", QVariant.Int),
                                QgsField("s_strata", QVariant.Double)])
        layer.updateFields()
        return layer

    def process_xml_file(self, xml_file, layer):
        try:
            # Попытка разобрать XML-файл
            tree = ET.parse(xml_file)
            root = tree.getroot()

            # Получить данные из XML
            path_area = float(root.find('.//_taxation/_path_area').text)
            strata, s_strata = self.assign_excel_values(xml_file)  # Извлекаем и название страты, и s_strata
            if s_strata is not None:
                s_strata = round(s_strata, 1)

            tax_info_list = []
            for tax_tree in root.findall('.//_taxation/_taxTrees/TaxTree'):
                coef = tax_tree.find('_coef').text
                species = tax_tree.find('_species').text
                if coef == "0":
                    coef = "+"
                tax_info_list.append(coef + species)
            tax_info = "".join(tax_info_list)
            lat = float(root.find('.//_gpsData/_lat').text)
            lon = float(root.find('.//_gpsData/_lon').text)

            # Подсчет количества деревьев
            tree_count = len(root.findall('.//_trees/Tree'))

            # Создать объект на слой
            provider = layer.dataProvider()
            feature = QgsFeature()
            feature.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(lon, lat)))
            feature.setAttributes([xml_file, path_area, tax_info, strata, tree_count, s_strata])
            if provider.addFeature(feature):
                print(f"Объект добавлен на слой: {feature.attributes()}")
            else:
                print(f"Не удалось добавить объект на слой: {feature.attributes()}")
        except ET.ParseError as e:
            # Вывод сообщения об ошибке разбора XML
            print(f"Ошибка при парсинге xml-файла {xml_file}: {e}")




    def create_layer_with_additional_info(self):
        # Создать новый векторный слой с дополнительной информацией
        layer = QgsVectorLayer("Point", "XML Деревья", "memory")
        if not layer.isValid():
            return None
        provider = layer.dataProvider()
        provider.addAttributes([QgsField("Путь xml-файла", QVariant.String),
                                
                                QgsField("ID", QVariant.String),
                                QgsField("Дата таксации", QVariant.String),
                                QgsField("Широта", QVariant.String),
                                QgsField("Долгота", QVariant.String),
                                QgsField("ВНУМ", QVariant.String),
                                QgsField("Субъект", QVariant.String),
                                QgsField("Лесничество", QVariant.String),
                                QgsField("Уч. лесничество", QVariant.String),
                                QgsField("Участок, урочище и т.п.", QVariant.String),
                                QgsField("Квартал", QVariant.String),
                                QgsField("Выдел", QVariant.String),
                                QgsField("Лп. выдел", QVariant.String),
                                QgsField("Площадь выдела", QVariant.String),
                                QgsField("Площадь лп. выдела", QVariant.String),
                                QgsField("Исполнитель", QVariant.String),
                                QgsField("Номер ППН", QVariant.String),

                                QgsField("ТХ / Непокрытые лесом (катег.)", QVariant.String),
                                QgsField("ТХ / Категория защитности", QVariant.String),
                                QgsField("ТХ / Состав", QVariant.String),
                                QgsField("ТХ / Ярус", QVariant.String),
                                QgsField("ТХ / Полнота", QVariant.String),
                                QgsField("ТХ / Бонитет", QVariant.String),
                                QgsField("ТХ / Тип леса", QVariant.String),
                                QgsField("ТХ / Запас , дес. м3/га", QVariant.String),
                                QgsField("ПП / Тип ПП", QVariant.String),
                                QgsField("ПП / Число деревьев", QVariant.String),
                                QgsField("Страта", QVariant.String),
                                QgsField("ЭЛ / Главная порода", QVariant.String),
                                QgsField("ЭЛ / Доля участия породы", QVariant.String),
                                QgsField("ЭЛ / Возраст", QVariant.String),
                                QgsField("ЭЛ / Высота", QVariant.String),
                                QgsField("ЭЛ / Диаметр", QVariant.String),
                                QgsField("ПО / Причина 1", QVariant.String),
                                QgsField("ПО / Причина 2", QVariant.String),
                                QgsField("ПО / Причина 3", QVariant.String),
                                QgsField("ПО / Причина 4", QVariant.String),
                                QgsField("Аренда (Да/Нет)", QVariant.String),
                                QgsField("Аренда (Примечание)", QVariant.String),

                                QgsField("Д / Номер", QVariant.String),
                                QgsField("Д / Порода", QVariant.String),
                                QgsField("Д / Диаметр", QVariant.String),
                                QgsField("Д / Высота", QVariant.String),
                                QgsField("Д / Категория состояния", QVariant.String),
                                QgsField("Д / Ярус", QVariant.String),
                                QgsField("Д / Класс урожайности", QVariant.String),

                                QgsField("Признак повреждения 1", QVariant.Int),
                                QgsField("Признак повреждения 2", QVariant.Int),
                                QgsField("Признак повреждения 3", QVariant.Int),
                                QgsField("Признак повреждения 4", QVariant.Int),
                                QgsField("Признак повреждения 5", QVariant.Int),
                                QgsField("Причина повреждения 1", QVariant.Int),
                                QgsField("Причина повреждения 2", QVariant.Int),
                                QgsField("Причина повреждения 3", QVariant.Int),
                                QgsField("Причина повреждения 4", QVariant.Int),
                                QgsField("Причина повреждения 5", QVariant.Int),

                                QgsField("Примечание (к дереву)", QVariant.String),
                                QgsField("Примечание (к выделу)", QVariant.String),
                                QgsField("GV", QVariant.String),
                                QgsField("Площадь ППН", QVariant.String),
                                QgsField("Площадь страты", QVariant.String),
                                QgsField("Количество живых деревьев", QVariant.String),
                                QgsField("Высота породы", QVariant.String),
                                QgsField("Радиус ППН", QVariant.String),

                                QgsField("h_vost1", QVariant.Double),
                                QgsField("v_tree", QVariant.Double)])
        layer.updateFields()
        return layer

    def process_xml_file_with_additional_info(self, xml_file, layer):

        try:
                
            # Попытка разобрать XML-файл
            tree = ET.parse(xml_file)
            root = tree.getroot()

            # Получить данные из XML
            path_area = float(root.find('.//_taxation/_path_area').text)

            strata = None
            s_strata = None
            #strata, s_strata = self.assign_excel_values(xml_file)  # Извлекаем и название страты, и s_strata
            if s_strata is not None:
                s_strata = round(s_strata, 1)

            tax_info_list = []
            for tax_tree in root.findall('.//_taxation/_taxTrees/TaxTree'):
                coef = tax_tree.find('_coef').text
                species = tax_tree.find('_species').text
                if coef == "0":
                    coef = "+"
                tax_info_list.append(coef + species)
            tax_info = "".join(tax_info_list)
            lat = float(root.find('.//_gpsData/_lat').text)
            lon = float(root.find('.//_gpsData/_lon').text)

            # Извлечение дополнительной информации
            tax_species = root.find('.//_taxation/_taxTrees/TaxTree/_species').text
            tax_coef = root.find('.//_taxation/_taxTrees/TaxTree/_coef').text
            tax_age = root.find('.//_taxation/_taxTrees/TaxTree/_age').text
            tax_h = root.find('.//_taxation/_taxTrees/TaxTree/_h').text
            tax_d = root.find('.//_taxation/_taxTrees/TaxTree/_d').text
            dam1 = root.find('.//_header/_mainDamages').text
            
            # Обработка всех элементов _Damages
            dam2_elements = root.findall('.//_header/_Damages[1]')
            dam2 = dam2_elements[0].text if dam2_elements else None
            
            dam3_elements = root.findall('.//_header/_Damages[2]')
            dam3 = dam3_elements[0].text if dam3_elements else None
            
            dam4_elements = root.findall('.//_header/_Damages[3]')
            dam4 = dam4_elements[0].text if dam4_elements else None

            arenda_date = root.find('.//_header/_tenantry/_tenantryTo').text
            lpt_date = root.find('.//_header/_gpsData/_lpt_date').text
            arenda = arenda_date >= lpt_date
            arenda_info = f"c {arenda_date} по {lpt_date}" if arenda else None

            r_ppn = root.find('.//_header/_plot/_r').text
            s_ppn = round((math.pi * (float(root.find('.//_header/_plot/_r').text) **2))/10000, 3)

            # Получить все деревья в XML-файле
            all_trees = root.findall('.//_trees/Tree')

            tree_count = len(all_trees)  # Подсчет количества деревьев
            # Подсчитать число деревьев с _stateCategory < 5
            LTrees = sum(1 for tree in all_trees if int(tree.find('_stateCategory').text) < 5)

            # Извлекаем необходимые данные из корневого элемента XML-файла
            file_id = root.find('.//_header/_id').text
            lpt_date = root.find('.//_header/_gpsData/_lpt_date').text
            lat = float(root.find('.//_header/_gpsData/_lat').text)
            lon = float(root.find('.//_header/_gpsData/_lon').text)
            altitude = float(root.find('.//_header/_gpsData/_altitude').text)
            region = root.find('.//_header/_region').text
            forestry = root.find('.//_header/_forestry').text
            forestry_area = root.find('.//_header/_forestryArea').text
            sub_forestry = root.find('.//_header/_subForestry').text
            kvartal = root.find('.//_header/_kvartal').text
            patch = root.find('.//_header/_patch').text
            forest_pathology_section = root.find('.//_header/_forestPatologySection').text
            path_area = float(root.find('.//_taxation/_path_area').text)
            forest_pathology_section_s = root.find('.//_header/_forestPatologySectionS').text
            worker = root.find('.//_header/_worker').text
            _n = root.find('.//_n')
            _n = _n.text if _n is not None else None

            _tax1 = root.find('.//_taxation/_notForestType')
            _tax1 = _tax1.text if _tax1 is not None else None
            _tax2 = root.find('.//_taxation/_landuse').text
            _tax3 = tax_info
            _tax4 = root.find('.//taxation/TaxTree/_layer')
            _tax4 = _tax4.text if _tax4 is not None else None
            _tax5 = root.find('.//_taxation/_density').text
            _tax6 = root.find('.//_taxation/_bonitet').text
            _tax7 = root.find('.//_taxation/_forest_type').text
            _tax8 = root.find('.//_taxation/_stock').text

            _pp = root.find('.//_header/_blankType').text
            _nTrees = tree_count
            _strata = root.find('.//_taxation/_strata').text
            if _strata is None:
                _strata = strata

            _el1 = tax_species
            _el2 = tax_coef
            _el3 = tax_age
            _el4 = tax_h
            _el5 = tax_d

            _dmg1 = dam1
            _dmg2 = dam2
            _dmg3 = dam3
            _dmg4 = dam4

            _arnd1 = arenda
            _arnd2 = arenda_info

            # Деревья (подеревная информация)

            _desc = root.find('.//_header/_description')
            _desc = _desc.text if _desc is not None else None
            _gv = root.find('.//_header/_gpsData/_gpsDataType').text
            _sPPN = s_ppn
            _sStrata = s_strata
            # Число живых деревьев
            # Высота породы
            _rPPN = r_ppn
    

            # Обработать каждое дерево Tree
            provider = layer.dataProvider()  # Переместим создание провайдера за пределы цикла
            for tree_element in root.findall('.//_trees/Tree'):
                # Получить номер и породу дерева
                tree_number = tree_element.find('_n').text
                tree_species = tree_element.find('_species').text
                tree_d = tree_element.find('_d').text
                tree_h = tree_element.find('_h').text
                tree_sc = int(tree_element.find('_stateCategory').text)
                tree_l = tree_element.find('_layer').text
                tree_pc = tree_element.find('_productivityClass').text

                desc = tree_element.find('.//_trees/Tree/_description')
                desc = desc.text if desc is not None else None
                
                # Найти такую же породу в разделе TaxTree
                tax_tree_species = root.find(f".//_taxation/_taxTrees/TaxTree[_species='{tree_species}']")
                # Если находим такую же породу, извлекаем значение высоты
                if tax_tree_species is not None:
                    _hSpecies = float(tax_tree_species.find('_h').text)
                else:
                    _hSpecies = None  # Или другое значение по умолчанию, если не найдено                
                
                # Извлечение признаков и причин для каждого дерева
                priznaki = []
                prichiny = []
                for item in tree_element.findall('_weakAppearance/item'):
                    priznak = item.find('key/int').text
                    prichina = item.find('value/int').text
                    priznaki.append(priznak)
                    prichiny.append(prichina)
                # Заполнение пустых значений, если их меньше 5
                priznaki.extend([None] * (5 - len(priznaki)))
                prichiny.extend([None] * (5 - len(prichiny)))

                # Высота восстановленная (по породам и ярусам по известным высотам)
                h_vost1 = self.calculate_average_height(root, tree_species, tree_l)
                if h_vost1 is not None:
                    h_vost1 = round(h_vost1,1)

                if tree_h is None:
                    tree_h = h_vost1

                tree_svet = ['ОС','Б','С','ДН','ОЛС','ИВ','Р','ОЛ','Л','Я','Д','ЯБ','ОР','ГШ']
                tree_ten = ['Е','КЛ','ЛП','ОЛЧ','В','Ч','К','П','ЛЩ','КРУ']
                if tree_species in tree_svet:
                    tree_svet_coef = 0.4
                elif tree_species in tree_ten:
                    tree_svet_coef = 0.42
                else:
                    tree_svet_coef = 0.41

                tree_g = math.pi/4*(int(tree_d)/100)**2
                # Проверяем, что tree_h не равен None, не равен 0 и не пустая строка
                if tree_h is not None and tree_h != 0 and str(tree_h).strip():
                    tree_v = (float(tree_h)+3)*tree_svet_coef*tree_g
                # Проверяем, что _hSpecies не равен None, не равен 0 и не пустая строка
                elif _hSpecies is not None and _hSpecies != 0 and str(_hSpecies).strip():
                    tree_v = (float(_hSpecies)+3)*tree_svet_coef*tree_g
                else:
                    tree_v = None
                    
                
                tree_v = round(tree_v,4)

                # Создать объект на слой для каждого дерева
                feature = QgsFeature()
                feature.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(lon, lat)))
                feature.setAttributes([
                    xml_file, 
                
                    file_id, lpt_date, lat, lon, altitude, region, forestry,
                    forestry_area, sub_forestry, kvartal, patch, forest_pathology_section,
                    path_area, forest_pathology_section_s, worker, _n,
                    _tax1, _tax2, _tax3, _tax4, _tax5, _tax6, _tax7, _tax8, 
                    _pp, _nTrees, _strata,
                    _el1, _el2, _el3, _el4, _el5,
                    _dmg1, _dmg2, _dmg3, _dmg4,
                    _arnd1, _arnd2,

                    tree_number, tree_species, tree_d, tree_h, tree_sc, tree_l, tree_pc,
                    *priznaki, *prichiny,
                    desc,

                    _desc, _gv, _sPPN, _sStrata, LTrees, _hSpecies, _rPPN,
                    
                    
                    h_vost1, tree_v
                ])

                # Попытка добавления объекта на слой
                if provider.addFeature(feature):
                    print(f"Объект добавлен на слой: {feature.attributes()}")
                else:
                    print(f"Не удалось добавить объект на слой: {feature.attributes()}")

            print(f"Обработка файла {xml_file} завершена")
        except ET.ParseError as e:
            # Вывод сообщения об ошибке разбора XML
            print(f"Ошибка при парсинге xml-файла {xml_file}: {e}")


    def calculate_average_height(self, root, species, layer):
        heights = []
        for tree_element in root.findall('.//_trees/Tree'):
            tree_species = tree_element.find('_species').text
            tree_layer = tree_element.find('_layer').text
            tree_height = tree_element.find('_h')
            if tree_species == species and tree_layer == layer and tree_height is not None and tree_height.text.strip() != '' and tree_height.text.strip() != '0':
                heights.append(float(tree_height.text))
        if heights:
            return sum(heights) / len(heights)
        else:
            return None










    def get_xml_files_in_folder(self, folder_path):
        # Получить список XML-файлов в указанной папке и всех ее подпапках
        xml_files = []
        for root_dir, _, files in os.walk (folder_path):
            for filename in files:
                if filename.endswith(".xml"):
                    xml_files.append(os.path.join(root_dir, filename))
        return xml_files

    def find_executable(self):
        # Проверяем, установлен ли WinRAR
        possible_winrar_paths = [
            os.path.join(os.environ["ProgramFiles"], "WinRAR", "WinRAR.exe"),
            os.path.join(os.environ["ProgramFiles(x86)"], "WinRAR", "WinRAR.exe"),
            os.path.join(os.environ["ProgramW6432"], "WinRAR", "WinRAR.exe"),
        ]

        for path in possible_winrar_paths:
            if os.path.exists(path):
                return path

        # Проверяем, установлен ли 7-Zip
        possible_7zip_paths = [
            os.path.join(os.environ["ProgramFiles"], "7-Zip", "7z.exe"),
            os.path.join(os.environ["ProgramFiles(x86)"], "7-Zip", "7z.exe"),
            os.path.join(os.environ["ProgramW6432"], "7-Zip", "7z.exe"),
        ]

        for path in possible_7zip_paths:
            if os.path.exists(path):
                return path

        # Добавляем пути для других архиваторов
        other_possible_paths = [
            # Путь к PeaZip
            os.path.join(os.environ["ProgramFiles"], "PeaZip", "peazip.exe"),
            os.path.join(os.environ["ProgramFiles(x86)"], "PeaZip", "peazip.exe"),
            os.path.join(os.environ["ProgramW6432"], "PeaZip", "peazip.exe"),
            # Другие пути к архиваторам...
        ]

        for path in other_possible_paths:
            if os.path.exists(path):
                return path

        # Если ни один архиватор не найден, предлагаем пользователю выбрать путь к исполняемому файлу
        return self.choose_executable_path()

    def choose_executable_path(self):
        dialog = QFileDialog()
        dialog.setFileMode(QFileDialog.ExistingFile)
        dialog.setNameFilter("Исполняемый файл (*.exe)")
        if dialog.exec_():
            return dialog.selectedFiles()[0]

        return None
    
    def get_xml_files_in_archive(self, archive_file):
        xml_files = []
        try:
            # Получаем путь к исполняемому файлу архиватора
            executable_path = self.find_executable()

            if not executable_path:
                raise ValueError("Не найден исполняемый файл для распаковки архива")
                
            # Создаем временную папку для извлечения файлов
            self.temp_dir = tempfile.mkdtemp()

            # Запускаем команду для распаковки архива
            command = [executable_path, 'x', archive_file, f'-o{self.temp_dir}']
            subprocess.run(command, check=True)

            # Проходим по всем извлеченным файлам
            for root_dir, _, files in os.walk(self.temp_dir):
                for filename in files:
                    if filename.endswith(".xml"):
                        xml_file_path = os.path.join(root_dir, filename)
                        # Добавляем файл в список только если он существует
                        if os.path.exists(xml_file_path):
                            xml_files.append(xml_file_path)
                        else:
                            print(f"Файл {xml_file_path} не найден")
        except Exception as e:
            print(f"Ошибка при разархивации {archive_file}: {e}")

        return xml_files
    

#ВРЕМЕННО определяем название страты по номеру ППН _n (если он установлен)
#Для привязки страты к ППН на этапе её добавления из xml-файлов
#Ест очень много времени, надо будет упростить
    def assign_excel_values(self, xml_file):
        try:
            # Открываем файл Excel
            excel_file_path = os.path.join(cmd_folder, "Приложение 1.xlsx")
            wb = openpyxl.load_workbook(excel_file_path)
            ws = wb.active
            
            # Получаем значение _n из XML-файла
            tree = ET.parse(xml_file)
            root = tree.getroot()
            _n_value = root.find('.//_n').text
            
            # Ищем значение _n в столбце K и получаем соответствующее значение из столбца L и M
            for row in ws.iter_rows(min_row=2, max_col=13, max_row=ws.max_row):  # Начинаем с 2 строки, так как первая строка - заголовки
                if row[10].value == _n_value:
                    excel_strata = row[11].value
                    excel_s_strata = row[12].value
                    return excel_strata, excel_s_strata

            # Если значение _n не найдено в файле Excel, выводим предупреждение
            print(f"ППН под номером {_n_value} не найдена в таблице страт (Приложение 1)")
            return None, None
        except Exception as e:
            print(f"Ошибка при присвоении значений из Excel: {e}")
            return None, None


#Делаем слой точек ППН сразу по таблицам страт (Приложение 1) (предполагаем, что он (excel файл с таблицей страт) есть у пользователя в исходных данных, в корне плагина)
#Работает почти мгновенно. Думаю, хорошая база и основа для создания замены метода по привязке страты (и её данных)
    def load_excel_data_to_map(self):
        # Проверяем, существует ли файл Excel с таблицей страт (Приложение 1)
        if not os.path.exists(self.excel_file_path):
            print(f"Excel файл '{self.excel_file_path}' не найден")
            return

        try:
            # Создаем слой для добавления точек
            layer = QgsVectorLayer("Point", "Excel Data", "memory")
            if not layer.isValid():
                print("Не удалось создать слой по данным Excel")
                return

            # Получаем провайдера данных слоя
            provider = layer.dataProvider()

            # Добавляем поля в атрибутивную таблицу слоя
            provider.addAttributes([QgsField("Latitude", QVariant.Double),
                                    QgsField("Longitude", QVariant.Double),
                                    QgsField("Area", QVariant.Double),  # Площадь страты
                                    QgsField("Name", QVariant.String),  # Название страты
            
                                    QgsField("Субъект", QVariant.String),  # Субъект
                                    QgsField("Зона ЛПУ", QVariant.String),  # Зона ЛПУ
                                    QgsField("Лесной район", QVariant.String),  # Лесной район
            
                                    QgsField("Лесничество", QVariant.String),  # Лесничество
                                    QgsField("Участковое лесничество", QVariant.String),  # Участковое лесничество
                                    QgsField("Лесной участок", QVariant.String),  # Лесной участок
                                    QgsField("KV", QVariant.Int),  # КВ
                                    QgsField("VYD", QVariant.Int),  # ВЫД
            
                                    QgsField("Номер ППН", QVariant.Int),  # Номер ППН
            
                                    QgsField("Статус", QVariant.String),  # Статус
                                    QgsField("Год статуса", QVariant.String)])  # Год статуса

            # Обновляем поля слоя
            layer.updateFields()

            current_year = datetime.now().year

            # Загружаем данные из Excel файла
            wb = load_workbook(filename=self.excel_file_path)
            ws = wb.active

            # Находим столбец с текущим годом
            column_index = None
            for cell in ws[6]:
                if cell.value == current_year:
                    column_index = cell.column
                    break

            if column_index is None:
                print(f"Столбец для текущего года '{current_year}' не найден")
                return

            # Читаем данные в Excel и добавляем точки на карту в QGIS
            for row in ws.iter_rows(min_row=2, max_col=13, max_row=ws.max_row):
                # Получаем координаты из Excel
                lat_excel_str = row[8].value
                lon_excel_str = row[9].value

                # Проверяем, что данные не пустые и являются числами
                if isinstance(lat_excel_str, (int, float)) and isinstance(lon_excel_str, (int, float)):
                    lat_excel = float(lat_excel_str)  # Колонка I (lat)
                    lon_excel = float(lon_excel_str)  # Колонка J (lon)

                    # Создаем точку на карте в QGIS
                    point = QgsPointXY(lon_excel, lat_excel)
                    feature = QgsFeature()
                    feature.setGeometry(QgsGeometry.fromPointXY(point))

                    # Получаем значения "Площадь" и "Название" страты из Excel, а также другие данные
                    area_value = row[12].value
                    if area_value is not None:
                        area_value = round(area_value,1)  # Колонка M
                    name_value = row[11].value  # Колонка L

                    #Регион, зона ЛПУ и лесной район
                    subject_value = row[0].value  # Колонка A
                    zonaLPU_value = row[1].value  # Колонка B
                    raion_value = row[2].value  # Колонка C

                    #Лесничество, уч. лесничество, участок/урочище, квартал, выдел
                    forestry_value = row[3].value  # Колонка D
                    uchforestry_value = row[4].value  # Колонка E
                    urochishe_value = row[5].value  # Колонка F
                    KV_value = row[6].value  # Колонка G
                    VYD_value = row[7].value  # Колонка H

                    #Номер ППН
                    nPPN_value = row[10].value  # Колонка K

                    # Получаем статус (Д - действует, Р - в резерве, С - списана) из столбца текущего года
                    status_value = ws.cell(row=row[0].row, column=column_index).value

                    # Ищем подходящее значение статуса в предыдущих столбцах
                    suitable_value_found = False
                    for col in reversed(range(13, column_index + 1)):
                        status_value = ws.cell(row=row[0].row, column=col).value
                        if status_value is not None and status_value.strip():
                            status_year = ws.cell(row=6, column=col).value
                            if status_year:
                                status_value = f"{status_value}"
                                status_year = f"{status_year}"
                                print(f"Выбран статус ППН: {status_value} на год {status_year}")
                                suitable_value_found = True
                                break

                    if not suitable_value_found:
                        print("За предыдущие года не найдено подходящего значения статуса")
                        print("Значение в ячейке слева:")
                        for col in reversed(range(13, column_index + 1)):
                            cell = ws.cell(row=row[0].row, column=col)
                            print(cell.value)

                    # Устанавливаем атрибуты для точки
                    feature.setAttributes([lat_excel, lon_excel, area_value, name_value, subject_value, zonaLPU_value, raion_value, forestry_value, uchforestry_value, urochishe_value, KV_value, VYD_value, nPPN_value, status_value, status_year])

                    # Добавляем точку в провайдер данных слоя
                    provider.addFeatures([feature])

            # Добавляем слой на карту в QGIS
            QgsProject.instance().addMapLayer(layer)
            self.layers.append(layer)

            print("Данные ППН из Excel успешно добавлены на карту")

        except Exception as e:
            print(f"Ошибка при добавлении данных на карту: {e}")