import os
from qgis.PyQt.QtWidgets import QAction, QFileDialog, QToolBar
from qgis.core import QgsVectorLayer, QgsField, QgsFeature, QgsGeometry, QgsPointXY, QgsProject
#from PyQt5.QtWidgets import QAction, QCheckBox
from PyQt5.QtCore import QVariant, QSettings
#from PyQt5.QtGui import QIcon
import xml.etree.ElementTree as ET
import tempfile
import shutil
import subprocess
import inspect
import openpyxl
import math
from datetime import datetime
from openpyxl import load_workbook

# Получение пути папки, содержащую текущий плагин
cmd_folder = os.path.split(inspect.getfile(inspect.currentframe()))[0]
# Временное название
_name = 'XML to GeoCSV'

class XML2CSVPlugin:
    def __init__(self, iface):
        self.iface = iface
        self.layers = []

        # Создание временной директории
        self.temp_dir = tempfile.mkdtemp()

        # Путь к файлу Excel с лесными стратами (+ районами, зонами ЛПУ и т.д.) (Приложение 1)
        self.excel_file_path = os.path.join(cmd_folder, "Приложение 1.xlsx")
        
    def initGui(self):
        # Создание панели инструментов
        self.toolbar = self.iface.addToolBar(_name+" панель инструментов")

        #Создание инструментов для открытия XML-файлов, папок и архивов
        self.actionOpenXmlFiles = QAction('Файл', self.iface.mainWindow())
        self.actionOpenFolders = QAction('Папка', self.iface.mainWindow())
        self.actionOpenArchives = QAction('Архив', self.iface.mainWindow())

        self.actionOpenXmlFiles.triggered.connect(self.open_xml_files)
        self.actionOpenFolders.triggered.connect(self.open_folders)
        self.actionOpenArchives.triggered.connect(self.open_archives)

        self.toolbar.addAction(self.actionOpenXmlFiles)
        self.toolbar.addAction(self.actionOpenFolders)
        self.toolbar.addAction(self.actionOpenArchives)

        self.actionOpenExcel = QAction('Страты', self.iface.mainWindow())
        self.actionOpenExcel.triggered.connect(self.load_excel_data_to_map)
        self.toolbar.addAction(self.actionOpenExcel)

        self.iface.addPluginToMenu("&"+_name, self.actionOpenXmlFiles)
        self.iface.addPluginToMenu("&"+_name, self.actionOpenFolders)
        self.iface.addPluginToMenu("&"+_name, self.actionOpenArchives)
        self.iface.addPluginToMenu("&"+_name, self.actionOpenExcel)

        self.iface.messageBar().pushMessage('Инструменты плагина '+_name+' загружены в панель инструментов')

    def unload(self):        
        # Удаление действий из панели инструментов при выгрузке плагина
        self.iface.removePluginMenu("&"+_name, self.actionOpenXmlFiles)
        self.iface.removePluginMenu("&"+_name, self.actionOpenFolders)
        self.iface.removePluginMenu("&"+_name, self.actionOpenArchives)
        self.iface.removePluginMenu("&"+_name, self.actionOpenExcel)

    def open_xml_files(self):
        # Открыть диалоговое окно для выбора XML-файлов
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("XML (*.xml)")
        if file_dialog.exec_():
            # Получить список выбранных файлов
            xml_files = file_dialog.selectedFiles()
            # Обработать выбранные файлы
            self.process_files(xml_files)

    def open_folders(self):
        # Открыть диалоговое окно для выбора папки
        folder_dialog = QFileDialog()
        folder_dialog.setFileMode(QFileDialog.Directory)
        folder_dialog.setOption(QFileDialog.ShowDirsOnly)
        folder_dialog.setAcceptMode(QFileDialog.AcceptOpen)
        if folder_dialog.exec_():
            # Получить путь к выбранной папке
            folder_path = folder_dialog.selectedFiles()[0]
            # Проверить, выбран ли архив или папка
            if folder_path.endswith('.rar') or folder_path.endswith('.7z') or folder_path.endswith('.zip'):
                # Если выбран архив, создать временную папку
                self.temp_dir = tempfile.mkdtemp()
                # Обработать архив
                xml_files = self.get_xml_files_in_archive(folder_path)
                # Удалить временную папку после обработки
                shutil.rmtree(self.temp_dir)
            else:
                # Если выбрана папка или файл, обработать её содержимое без использования временной папки
                if os.path.isdir(folder_path):
                    xml_files = self.get_xml_files_in_folder(folder_path)
                else:
                    xml_files = [folder_path]
            # Обработать XML-файлы
            self.process_files(xml_files)

    def open_archives(self):
        # Открыть диалоговое окно для выбора архивов
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("Архив (*.zip *.rar *.7z)")
        if file_dialog.exec_():
            # Получить список выбранных архивов
            archive_files = file_dialog.selectedFiles()
            for archive_file in archive_files:
                # Получить список XML-файлов в архиве
                xml_files = self.get_xml_files_in_archive(archive_file)
                # Обработать XML-файлы
                self.process_files(xml_files)

    def process_files(self, xml_files):
        if not xml_files:
            return

        # Создать новый слой
        layer = self.create_layer()
        if layer is None:
            print("Не удалось создать слой")
            return

        # Обработать каждый XML-файл и добавить объекты на слой
        for xml_file in xml_files:
            self.process_xml_file(xml_file, layer)

        # Добавить слой в проект
        self.layers.append(layer)
        QgsProject.instance().addMapLayer(layer)

    def create_layer(self):
        # Создать новый векторный слой
        layer = QgsVectorLayer("Point", "XML Layer", "memory")
        if not layer.isValid():
            return None
        provider = layer.dataProvider()
        provider.addAttributes([QgsField("xml_path", QVariant.String),
                                QgsField("path_area", QVariant.Double),
                                QgsField("tax_info", QVariant.String),
                                QgsField("strata", QVariant.String),
                                QgsField("tree_count", QVariant.Int)])
        layer.updateFields()
        return layer

    def process_xml_file(self, xml_file, layer):
        try:
            # Попытка разобрать XML-файл
            tree = ET.parse(xml_file)
            root = tree.getroot()

            # Получить данные из XML
            path_area = float(root.find('.//_taxation/_path_area').text)
            strata = self.assign_excel_values(xml_file)

            tax_info_list = []
            for tax_tree in root.findall('.//_taxation/_taxTrees/TaxTree'):
                coef = tax_tree.find('_coef').text
                species = tax_tree.find('_species').text
                if coef == "0":
                    coef = "+"
                tax_info_list.append(coef + species)
            tax_info = "".join(tax_info_list)
            lat = float(root.find('.//_gpsData/_lat').text)
            lon = float(root.find('.//_gpsData/_lon').text)

            # Подсчет количества деревьев
            tree_count = len(root.findall('.//_trees/Tree'))

            # Создать объект на слой
            provider = layer.dataProvider()
            feature = QgsFeature()
            feature.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(lon, lat)))
            feature.setAttributes([xml_file, path_area, tax_info, strata, tree_count])
            provider.addFeature(feature)
        except ET.ParseError as e:
            # Вывод сообщения об ошибке разбора XML
            print(f"Ошибка при парсинге xml-файла {xml_file}: {e}")


    def get_xml_files_in_folder(self, folder_path):
        # Получить список XML-файлов в указанной папке и всех ее подпапках
        xml_files = []
        for root_dir, _, files in os.walk (folder_path):
            for filename in files:
                if filename.endswith(".xml"):
                    xml_files.append(os.path.join(root_dir, filename))
        return xml_files

    def find_executable(self):
        # Проверяем, установлен ли WinRAR
        possible_winrar_paths = [
            os.path.join(os.environ["ProgramFiles"], "WinRAR", "WinRAR.exe"),
            os.path.join(os.environ["ProgramFiles(x86)"], "WinRAR", "WinRAR.exe"),
            os.path.join(os.environ["ProgramW6432"], "WinRAR", "WinRAR.exe"),
        ]

        for path in possible_winrar_paths:
            if os.path.exists(path):
                return path

        # Проверяем, установлен ли 7-Zip
        possible_7zip_paths = [
            os.path.join(os.environ["ProgramFiles"], "7-Zip", "7z.exe"),
            os.path.join(os.environ["ProgramFiles(x86)"], "7-Zip", "7z.exe"),
            os.path.join(os.environ["ProgramW6432"], "7-Zip", "7z.exe"),
        ]

        for path in possible_7zip_paths:
            if os.path.exists(path):
                return path

        # Добавляем пути для других архиваторов
        other_possible_paths = [
            # Путь к PeaZip
            os.path.join(os.environ["ProgramFiles"], "PeaZip", "peazip.exe"),
            os.path.join(os.environ["ProgramFiles(x86)"], "PeaZip", "peazip.exe"),
            os.path.join(os.environ["ProgramW6432"], "PeaZip", "peazip.exe"),
            # Другие пути к архиваторам...
        ]

        for path in other_possible_paths:
            if os.path.exists(path):
                return path

        # Если ни один архиватор не найден, предлагаем пользователю выбрать путь к исполняемому файлу
        return self.choose_executable_path()

    def choose_executable_path(self):
        dialog = QFileDialog()
        dialog.setFileMode(QFileDialog.ExistingFile)
        dialog.setNameFilter("Исполняемый файл (*.exe)")
        if dialog.exec_():
            return dialog.selectedFiles()[0]

        return None
    
    def get_xml_files_in_archive(self, archive_file):
        xml_files = []
        try:
            # Получаем путь к исполняемому файлу архиватора
            executable_path = self.find_executable()

            if not executable_path:
                raise ValueError("Не найден исполняемый файл для распаковки архива")
                
            # Создаем временную папку для извлечения файлов
            self.temp_dir = tempfile.mkdtemp()

            # Запускаем команду для распаковки архива
            command = [executable_path, 'x', archive_file, f'-o{self.temp_dir}']
            subprocess.run(command, check=True)

            # Проходим по всем извлеченным файлам
            for root_dir, _, files in os.walk(self.temp_dir):
                for filename in files:
                    if filename.endswith(".xml"):
                        xml_file_path = os.path.join(root_dir, filename)
                        # Добавляем файл в список только если он существует
                        if os.path.exists(xml_file_path):
                            xml_files.append(xml_file_path)
                        else:
                            print(f"Файл {xml_file_path} не найден")
        except Exception as e:
            print(f"Ошибка при разархивации {archive_file}: {e}")

        return xml_files
    

#ВРЕМЕННО определяем название страты по номеру ППН _n (если он установлен)
#Для привязки страты к ППН на этапе её добавления из xml-файлов
#Ест очень много времени, надо будет упростить
    def assign_excel_values(self, xml_file):
        try:
            # Открываем файл Excel
            excel_file_path = os.path.join(cmd_folder, "Приложение 1.xlsx")
            wb = openpyxl.load_workbook(excel_file_path)
            ws = wb.active
            
            # Получаем значение _n из XML-файла
            tree = ET.parse(xml_file)
            root = tree.getroot()
            _n_value = root.find('.//_n').text
            
            # Ищем значение _n в столбце K и получаем соответствующее значение из столбца L
            for row in ws.iter_rows(min_row=2, max_col=12, max_row=ws.max_row):  # Начинаем с 2 строки, так как первая строка - заголовки
                if row[10].value == _n_value:
                    excel_value = row[11].value
                    return excel_value

            # Если значение _n не найдено в файле Excel, выводим предупреждение
            print(f"ППН под номером {_n_value} не найдена в таблице страт (Приложение 1)")
            return None
        except Exception as e:
            print(f"Ошибка при присвоении значений из Excel: {e}")
            return None


#Делаем слой точек ППН сразу по таблицам страт (Приложение 1) (предполагаем, что он (excel файл с таблицей страт) есть у пользователя в исходных данных, в корне плагина)
#Работает почти мгновенно. Думаю, хорошая база и основа для создания замены метода по привязке страты (и её данных)
    def load_excel_data_to_map(self):
        # Проверяем, существует ли файл Excel с таблицей страт (Приложение 1)
        if not os.path.exists(self.excel_file_path):
            print(f"Excel файл '{self.excel_file_path}' не найден")
            return

        try:
            # Создаем слой для добавления точек
            layer = QgsVectorLayer("Point", "Excel Data", "memory")
            if not layer.isValid():
                print("Не удалось создать слой по данным Excel")
                return

            # Получаем провайдера данных слоя
            provider = layer.dataProvider()

            # Добавляем поля в атрибутивную таблицу слоя
            provider.addAttributes([QgsField("Latitude", QVariant.Double),
                                    QgsField("Longitude", QVariant.Double),
                                    QgsField("Area", QVariant.Double),  # Площадь страты
                                    QgsField("Name", QVariant.String),  # Название страты
            
                                    QgsField("Субъект", QVariant.String),  # Субъект
                                    QgsField("Зона ЛПУ", QVariant.String),  # Зона ЛПУ
                                    QgsField("Лесной район", QVariant.String),  # Лесной район
            
                                    QgsField("Лесничество", QVariant.String),  # Лесничество
                                    QgsField("Участковое лесничество", QVariant.String),  # Участковое лесничество
                                    QgsField("Лесной участок", QVariant.String),  # Лесной участок
                                    QgsField("KV", QVariant.Int),  # КВ
                                    QgsField("VYD", QVariant.Int),  # ВЫД
            
                                    QgsField("Номер ППН", QVariant.Int),  # Номер ППН
            
                                    QgsField("Статус", QVariant.String),  # Статус
                                    QgsField("Год статуса", QVariant.String)])  # Год статуса

            # Обновляем поля слоя
            layer.updateFields()

            current_year = datetime.now().year

            # Загружаем данные из Excel файла
            wb = load_workbook(filename=self.excel_file_path)
            ws = wb.active

            # Находим столбец с текущим годом
            column_index = None
            for cell in ws[6]:
                if cell.value == current_year:
                    column_index = cell.column
                    break

            if column_index is None:
                print(f"Столбец для текущего года '{current_year}' не найден")
                return

            # Читаем данные в Excel и добавляем точки на карту в QGIS
            for row in ws.iter_rows(min_row=2, max_col=13, max_row=ws.max_row):
                # Получаем координаты из Excel
                lat_excel_str = row[8].value
                lon_excel_str = row[9].value

                # Проверяем, что данные не пустые и являются числами
                if isinstance(lat_excel_str, (int, float)) and isinstance(lon_excel_str, (int, float)):
                    lat_excel = float(lat_excel_str)  # Колонка I (lat)
                    lon_excel = float(lon_excel_str)  # Колонка J (lon)

                    # Создаем точку на карте в QGIS
                    point = QgsPointXY(lon_excel, lat_excel)
                    feature = QgsFeature()
                    feature.setGeometry(QgsGeometry.fromPointXY(point))

                    # Получаем значения "Площадь" и "Название" страты из Excel, а также другие данные
                    area_value = row[12].value  # Колонка M
                    name_value = row[11].value  # Колонка L

                    #Регион, зона ЛПУ и лесной район
                    subject_value = row[0].value  # Колонка A
                    zonaLPU_value = row[1].value  # Колонка B
                    raion_value = row[2].value  # Колонка C

                    #Лесничество, уч. лесничество, участок/урочище, квартал, выдел
                    forestry_value = row[3].value  # Колонка D
                    uchforestry_value = row[4].value  # Колонка E
                    urochishe_value = row[5].value  # Колонка F
                    KV_value = row[6].value  # Колонка G
                    VYD_value = row[7].value  # Колонка H

                    #Номер ППН
                    nPPN_value = row[10].value  # Колонка K

                    # Получаем статус (Д - действует, Р - в резерве, С - списана) из столбца текущего года
                    status_value = ws.cell(row=row[0].row, column=column_index).value

                    # Ищем подходящее значение статуса в предыдущих столбцах
                    suitable_value_found = False
                    for col in reversed(range(13, column_index + 1)):
                        status_value = ws.cell(row=row[0].row, column=col).value
                        if status_value is not None and status_value.strip():
                            status_year = ws.cell(row=6, column=col).value
                            if status_year:
                                status_value = f"{status_value}"
                                status_year = f"{status_year}"
                                print(f"Выбран статус ППН: {status_value} на год {status_year}")
                                suitable_value_found = True
                                break

                    if not suitable_value_found:
                        print("За предыдущие года не найдено подходящего значения статуса")
                        print("Значение в ячейке слева:")
                        for col in reversed(range(13, column_index + 1)):
                            cell = ws.cell(row=row[0].row, column=col)
                            print(cell.value)

                    # Устанавливаем атрибуты для точки
                    feature.setAttributes([lat_excel, lon_excel, area_value, name_value, subject_value, zonaLPU_value, raion_value, forestry_value, uchforestry_value, urochishe_value, KV_value, VYD_value, nPPN_value, status_value, status_year])

                    # Добавляем точку в провайдер данных слоя
                    provider.addFeatures([feature])

            # Добавляем слой на карту в QGIS
            QgsProject.instance().addMapLayer(layer)
            self.layers.append(layer)

            print("Данные ППН из Excel успешно добавлены на карту")

        except Exception as e:
            print(f"Ошибка при добавлении данных на карту: {e}")